from openpyxl import load_workbook

from stockroom.altium.datasource import ALTIUM_COLUMNS, emit_xlsx, row_for
from stockroom.model.part import AltiumRef, Datasheet, PartRecord, Purchase


def _part():
    return PartRecord(
        id="bq24074rgtt", display_name="BQ24074 Charger", category="ICs",
        mpn="BQ24074RGTT", manufacturer="Texas Instruments", value="BQ24074RGTT",
        description="Li-Ion charger, VQFN-16",
        datasheet=Datasheet(source_url="https://ti.com/ds.pdf"),
        purchase=[Purchase(vendor="DigiKey", part_number="296-1", url="https://dk/1", stock=42)],
        altium_symbol=AltiumRef(lib="BQ24074RGTT.SchLib", name="BQ24074RGTT"),
        altium_footprint=AltiumRef(lib="BQ24074RGTT.PcbLib", name="VQFN-16"),
    )


def test_row_maps_reserved_and_field_columns():
    row = row_for(_part())
    assert row["MPN"] == "BQ24074RGTT"
    assert row["Library Ref"] == "BQ24074RGTT"
    assert row["Library Path"] == "BQ24074RGTT.SchLib"
    assert row["Footprint Ref"] == "VQFN-16"
    assert row["Footprint Path"] == "BQ24074RGTT.PcbLib"
    assert row["Value"] == "BQ24074RGTT"
    assert row["Manufacturer"] == "Texas Instruments"
    assert row["Description"] == "Li-Ion charger, VQFN-16"
    assert row["ComponentLink1URL"] == "https://ti.com/ds.pdf"
    assert row["ComponentLink1Description"] == "Datasheet"
    assert row["SupplierPartNumber"] == "296-1"
    assert row["Stock"] == "42"


def test_emit_writes_header_plus_one_row_sorted(tmp_path):
    out = tmp_path / "stockroom-parts.xlsx"
    n = emit_xlsx([_part()], out)
    assert n == 1
    wb = load_workbook(out)
    ws = wb["Parts"]
    assert [c.value for c in ws[1]] == ALTIUM_COLUMNS
    assert ws.cell(row=2, column=1 + ALTIUM_COLUMNS.index("MPN")).value == "BQ24074RGTT"
