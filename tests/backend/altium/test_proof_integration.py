"""End-to-end proof of the backend pipeline over a real Altium library: attach an .IntLib
(extract -> store loose -> read names -> set refs) then regenerate the DbLib + data source and
assert the emitted row + mappings. Uses the committed S1M sample so it runs with no external
downloads; the real 3-proof-part run is scripts/altium_proof.py (owner's assets)."""
from pathlib import Path

from openpyxl import load_workbook

from stockroom.model.part import PartRecord

FIX = Path(__file__).parent / "fixtures"


def test_attach_intlib_then_regenerate_lands_a_row(library_ops):
    ops = library_ops
    ops.lib.parts_dir.mkdir(parents=True, exist_ok=True)
    (ops.lib.parts_dir / "s1m.json").write_text(
        PartRecord(
            id="s1m", display_name="S1M diode", category="Diodes", mpn="S1M",
            manufacturer="Multicomp", description="1A SMA rectifier", value="S1M",
        ).dumps(),
        encoding="utf-8",
    )
    ops.attach_altium_assets("s1m", FIX / "sample.IntLib")

    result = ops.regenerate_altium_dblib()

    assert result["emitted"] == 1
    ws = load_workbook(result["xlsx"])["Parts"]
    row = {
        ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value
        for c in range(1, ws.max_column + 1)
    }
    assert row["MPN"] == "S1M"
    assert row["Value"] == "S1M"
    assert row["Manufacturer"] == "Multicomp"
    assert row["Library Path"] == "s1m.SchLib"
    assert row["Library Ref"] == "S1M"
    assert row["Footprint Path"] == "s1m.PcbLib"
    assert row["Footprint Ref"] == "DIOM5227X270N"

    dblib = result["dblib"].read_text(encoding="utf-8")
    assert "TableName=Parts$" in dblib
    assert "ParameterName=[Library Ref]" in dblib
    assert "ParameterName=[Footprint Path]" in dblib
    assert "DatabasePathRelative=1" in dblib
