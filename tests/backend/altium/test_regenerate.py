from openpyxl import load_workbook

from stockroom.model.part import AltiumRef, PartRecord


def _place_ready(pid, mpn):
    return PartRecord(
        id=pid, display_name=pid, category="ICs", mpn=mpn, manufacturer="TI",
        description="d", value=mpn,
        altium_symbol=AltiumRef(lib=f"{pid}.SchLib", name=mpn),
        altium_footprint=AltiumRef(lib=f"{pid}.PcbLib", name="FP"),
    )


def test_regenerate_emits_only_place_ready(library_ops):
    ops = library_ops
    ops.lib.parts_dir.mkdir(parents=True, exist_ok=True)
    (ops.lib.parts_dir / "a.json").write_text(_place_ready("a", "AAA").dumps(), encoding="utf-8")
    # not place-ready: no altium refs
    (ops.lib.parts_dir / "b.json").write_text(
        PartRecord(id="b", display_name="b", category="ICs", mpn="BBB").dumps(), encoding="utf-8")

    result = ops.regenerate_altium_dblib()

    assert result["emitted"] == 1
    assert "b" in result["skipped"]
    assert result["dblib"].exists()
    assert result["xlsx"].exists()
    ws = load_workbook(result["xlsx"])["Parts"]
    mpns = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert mpns == ["AAA"]

    # the derived .xlsx is ignored by a local .gitignore that travels with the library
    gi = result["dblib"].parent / ".gitignore"
    assert gi.exists() and "stockroom-parts.xlsx" in gi.read_text(encoding="utf-8")


def test_place_ready_does_not_require_a_persisted_value(library_ops):
    """The place-ready gate must NOT require record.value (nothing in the real pipeline sets it);
    the Value column is derived at emit time. A part with value="" but full identity + assets is
    emitted, with Value derived (an active's MPN)."""
    ops = library_ops
    ops.lib.parts_dir.mkdir(parents=True, exist_ok=True)
    rec = PartRecord(
        id="c", display_name="c", category="ICs", mpn="CCC", manufacturer="TI",
        description="a chip", value="",  # NOT populated by the real pipeline
        altium_symbol=AltiumRef(lib="c.SchLib", name="CCC"),
        altium_footprint=AltiumRef(lib="c.PcbLib", name="FP"),
    )
    (ops.lib.parts_dir / "c.json").write_text(rec.dumps(), encoding="utf-8")

    result = ops.regenerate_altium_dblib()

    assert result["emitted"] == 1  # emitted despite value==""
    from openpyxl import load_workbook
    ws = load_workbook(result["xlsx"])["Parts"]
    row = {ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)}
    assert row["Value"] == "CCC"  # derived (active -> MPN), not blank


def test_regenerate_is_idempotent(library_ops):
    """Regenerate runs on every data refresh; an unchanged .DbLib must not crash on an
    empty commit or spawn a noisy empty commit (GitRepo.commit no-ops on an empty diff)."""
    ops = library_ops
    ops.lib.parts_dir.mkdir(parents=True, exist_ok=True)
    (ops.lib.parts_dir / "a.json").write_text(_place_ready("a", "AAA").dumps(), encoding="utf-8")

    first = ops.regenerate_altium_dblib()
    head_after_first = ops.repo.head()
    second = ops.regenerate_altium_dblib()  # identical content

    assert second["emitted"] == 1
    assert ops.repo.head() == head_after_first  # no new (empty) commit
    assert second["dblib"].exists()
